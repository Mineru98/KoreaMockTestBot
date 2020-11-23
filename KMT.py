import io
import sys
import time
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup

# time setting
now = datetime.now()

# chrome setting
_chrome_options = webdriver.ChromeOptions() 
_chrome_options.add_argument('disable-infobars')
_chrome_options.add_argument('headless')
_chrome_options.add_argument("lang=ko_KR")
_chrome_options.add_argument('--no-sandbox')
_chrome_options.add_argument('--disable-dev-shm-usage')
driver = webdriver.Chrome('./chromedriver',options=_chrome_options)

# searching bot
def search_start(year=(now.year),month=(now.month),_class=3):
    year = int(year)
    month = int(month)
    _class = int(_class)
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    fileName = '%d년학도 %02d월 모의고사 등급컷 고등학교 %d학년.xlsx' % (year+1, month, _class)
    url = 'http://www.ebsi.co.kr/ebs/ent/entd/retrieveTrlNaatLiveGrdCutlnNw2017.ebs?examCd=%dN%02d&type=FULL_SUB&stdntGrd=%d' % (year,month,_class)
    driver.get(url)
    page = 1
    while True:
        try:
            path = '/html/body/div/div/div[3]/div[1]/ul/li[%d]/a' % page
            page = page + 1
            
            driver.find_element_by_xpath(path).click()
            element = WebDriverWait(driver,3).until(EC.presence_of_element_located((By.XPATH,'//*[@id="titles"]/th[2]')))
            time.sleep(0.01)
            
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')

            # 과목 갯수 구하기
            counter = soup.html.thead.find("tr", {"id":"titles"})
            test = counter.text[1:-1].replace('\n',',').replace('등급,','')
            wordAll=test.split(',')
            subject_count = 0
            
            # 타이틀 가져오기
            title = soup.html.body.find("div",{"class":"tabs full_sub"}).ul
            title = title.text.replace('\t','').replace('\n\n','\n').replace('\n\n','\n').replace('\n\n','\n')[1:-1].replace('\n',',').replace('/','').split(',')
            sheet.title = title[page-2]
                
            for subject in wordAll:
                sheet.cell(row=1, column=(2+3*subject_count)).value = subject
                subject_count = subject_count + 1
            
            sheet['A1'] = '등급'
            # count 수에 따른 원 표 백 배치
            # count 2이면 row=2 , column=2,5,8,
            for i in range(0,subject_count):
                sheet.cell(row=2, column=2+i*3).value='원점수'
                sheet.cell(row=2, column=3+i*3).value='표준점수'
                sheet.cell(row=2, column=4+i*3).value='백분위'
            
            step1 = soup.html.tbody.tr
            result = step1.find('td',{'class':''})
            for i in range(0, 8):
                try:
                    excel_title = "%d 등급" % (i+1)
                    sheet.cell(row=i+3, column=1).value = excel_title
                    stact = 0
                    while True:
                        try:
                            result = result.next_sibling
                            if result is not None:
                                sheet.cell(row=(3+i), column=(2+stact)).value = result.text
                                stact = stact + 1
                                
                        except AttributeError:
                            break
                    step1 = step1.next_sibling
                    result = step1.find('td',{'class':''})
                except AttributeError:
                    break

            
            ca1= sheet['A1']
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1= sheet['B1']
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1= sheet['E1']
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1= sheet['H1']
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1= sheet['K1']
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1= sheet['N1']
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            sheet.merge_cells('B1:D1')
            sheet.merge_cells('E1:G1')
            sheet.merge_cells('H1:J1')
            sheet.merge_cells('K1:M1')
            sheet.merge_cells('N1:P1')
            sheet.merge_cells('A1:A2')

             # 새로운 시트 생성
            sheet = wb.create_sheet('Next Sheet')
        except NoSuchElementException:
            ca1= sheet['A1']
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1= sheet['B1']
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1= sheet['E1']
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1= sheet['H1']
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1= sheet['K1']
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1= sheet['N1']
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            sheet.merge_cells('B1:D1')
            sheet.merge_cells('E1:G1')
            sheet.merge_cells('H1:J1')
            sheet.merge_cells('K1:M1')
            sheet.merge_cells('N1:P1')
            sheet.merge_cells('A1:A2')
            wb.remove(wb['Next Sheet'])
            wb.save(fileName)
            break

if len(sys.argv) == 1:
    search_start()
    driver.quit()
else:
    sys.stdout = io.TextIOWrapper((sys.stdout.detach()), encoding='utf8', newline='')
    sys.stderr = io.TextIOWrapper((sys.stderr.detach()), encoding='utf8', newline='')
    search_start(sys.argv[1],sys.argv[2],sys.argv[3])
    driver.quit()
